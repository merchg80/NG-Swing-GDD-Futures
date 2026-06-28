import argparse
import os
from pathlib import Path
from typing import Dict, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


PRODUCT_FILTER = "NG Swing GDD Futures"

REQUIRED_COLUMNS = {
    "trade_date": ["trade date", "tradedate", "trade_date"],
    "hub": ["hub"],
    "product": ["product"],
    "strip": ["strip"],
    "settlement_price": [
        "settlement price",
        "settlementprice",
        "settlement_price",
        "settle",
        "settle price",
        "settlement",
    ],
}


def clean_text(value) -> str:
    """
    Normalizes text for header matching.
    """
    if pd.isna(value):
        return ""

    text = str(value).strip().lower()
    text = text.replace("\n", " ")
    text = text.replace("\r", " ")
    text = " ".join(text.split())
    return text


def compact_text(value) -> str:
    """
    More aggressive normalization for header matching.
    """
    text = clean_text(value)
    for ch in [" ", "_", "-", ".", "/", "\\"]:
        text = text.replace(ch, "")
    return text


def normalize_date(value) -> Optional[pd.Timestamp]:
    """
    Converts Excel dates, Python dates, and date-like strings into a normalized pandas Timestamp.
    Returns None if the value cannot be parsed.
    """
    if pd.isna(value):
        return None

    parsed = pd.to_datetime(value, errors="coerce")

    if pd.isna(parsed):
        return None

    return pd.Timestamp(parsed).normalize()


def normalize_date_header(value) -> Optional[str]:
    """
    Normalizes a date value into YYYY-MM-DD.
    """
    parsed = normalize_date(value)

    if parsed is None:
        if value is None:
            return None
        value_text = str(value).strip()
        return value_text if value_text else None

    return parsed.strftime("%Y-%m-%d")


def header_matches(cell_value, aliases) -> bool:
    """
    Returns True if a cell value matches one of the allowed header aliases.
    """
    cleaned = clean_text(cell_value)
    compact = compact_text(cell_value)

    for alias in aliases:
        alias_clean = clean_text(alias)
        alias_compact = compact_text(alias)

        if cleaned == alias_clean:
            return True

        if compact == alias_compact:
            return True

    return False


def detect_header_row_and_columns(raw_no_header: pd.DataFrame) -> Tuple[int, Dict[str, int]]:
    """
    Finds the actual ICE table header row even if the data does not start on row 1.

    Looks for a row containing:
    TRADE DATE, HUB, PRODUCT, STRIP, SETTLEMENT PRICE

    Returns:
    header_row_index, column_map
    """

    best_row = None
    best_map = {}
    best_score = 0

    for row_idx in range(len(raw_no_header)):
        row_values = list(raw_no_header.iloc[row_idx, :])

        col_map = {}

        for col_idx, value in enumerate(row_values):
            for logical_name, aliases in REQUIRED_COLUMNS.items():
                if logical_name in col_map:
                    continue

                if header_matches(value, aliases):
                    col_map[logical_name] = col_idx

        score = len(col_map)

        if score > best_score:
            best_score = score
            best_row = row_idx
            best_map = col_map

        if all(key in col_map for key in REQUIRED_COLUMNS):
            print(f"Detected header row at Excel row {row_idx + 1}")
            print(f"Detected columns: {col_map}")
            return row_idx, col_map

    missing = [key for key in REQUIRED_COLUMNS if key not in best_map]

    raise ValueError(
        "Could not detect the ICE table header row.\n"
        f"Best detected row: {best_row + 1 if best_row is not None else None}\n"
        f"Best detected columns: {best_map}\n"
        f"Missing required columns: {missing}\n\n"
        "The script needs to find headers equivalent to: "
        "TRADE DATE, HUB, PRODUCT, STRIP, SETTLEMENT PRICE."
    )


def save_detected_table_preview(df: pd.DataFrame, audit_dir: str):
    """
    Saves a CSV preview of the detected table so you can confirm GitHub Actions
    is reading the correct rows and columns.
    """

    audit_path = Path(audit_dir)
    audit_path.mkdir(parents=True, exist_ok=True)

    preview_file = audit_path / "last_detected_table_preview.csv"

    preview_cols = [
        "TradeDateRaw",
        "Hub",
        "Product",
        "StripRaw",
        "DateHeader",
        "SettlementPrice",
    ]

    df[preview_cols].head(200).to_csv(preview_file, index=False)

    print(f"Detected table preview saved to: {preview_file}")


def read_daily_file(daily_file: str, date_field: str, audit_dir: str) -> pd.DataFrame:
    """
    Reads the ICE worksheet without assuming the data starts on the same row every day.

    The script:
    1. Reads the worksheet with no header.
    2. Searches for the actual table header row.
    3. Maps the needed columns by header name.
    4. Reads the rows below that header.
    """

    if not os.path.exists(daily_file):
        raise FileNotFoundError(f"Daily file not found: {daily_file}")

    raw_no_header = pd.read_excel(daily_file, header=None, dtype=object)

    if raw_no_header.empty:
        raise ValueError(f"The daily file appears empty: {daily_file}")

    header_row_idx, col_map = detect_header_row_and_columns(raw_no_header)

    data = raw_no_header.iloc[header_row_idx + 1:, :].copy()

    if data.empty:
        raise ValueError("The detected header row was found, but no data exists below it.")

    if date_field == "trade_date":
        selected_date_col = col_map["trade_date"]
        selected_date_label = "TRADE DATE"
    else:
        selected_date_col = col_map["strip"]
        selected_date_label = "STRIP"

    df = pd.DataFrame({
        "SelectedDateRaw": data.iloc[:, selected_date_col],
        "TradeDateRaw": data.iloc[:, col_map["trade_date"]],
        "StripRaw": data.iloc[:, col_map["strip"]],
        "Hub": data.iloc[:, col_map["hub"]],
        "Product": data.iloc[:, col_map["product"]],
        "SettlementPrice": data.iloc[:, col_map["settlement_price"]],
    })

    df = df.dropna(how="all")

    df["Hub"] = df["Hub"].astype(str).str.strip()
    df["Product"] = df["Product"].astype(str).str.strip()
    df["SelectedDate"] = df["SelectedDateRaw"].apply(normalize_date)

    df["DateHeader"] = df["SelectedDate"].apply(
        lambda x: x.strftime("%Y-%m-%d") if pd.notna(x) else None
    )

    df["SettlementPrice"] = pd.to_numeric(df["SettlementPrice"], errors="coerce")

    df = df[
        (df["Hub"].notna())
        & (df["Hub"].astype(str).str.strip() != "")
        & (df["Hub"].astype(str).str.lower() != "nan")
        & (df["Product"].notna())
        & (df["Product"].astype(str).str.strip() != "")
        & (df["Product"].astype(str).str.lower() != "nan")
    ].copy()

    print(f"Using date field: {selected_date_label}")
    print(f"Rows after header detection and cleanup: {len(df)}")

    save_detected_table_preview(df, audit_dir)

    return df


def print_diagnostics(df: pd.DataFrame):
    """
    Prints useful diagnostics to the GitHub Actions log.
    """

    product_matches = df[
        df["Product"].str.contains(PRODUCT_FILTER, case=False, na=False)
    ].copy()

    print("")
    print("========== DIAGNOSTICS ==========")
    print(f"Total detected data rows: {len(df)}")
    print(f"Rows where Product contains '{PRODUCT_FILTER}': {len(product_matches)}")

    if product_matches.empty:
        print("")
        print("No matching product rows found.")
        print("Sample Product values:")
        print(df["Product"].dropna().drop_duplicates().head(50).to_string(index=False))
        print("=================================")
        return

    available_dates = sorted(
        d for d in product_matches["DateHeader"].dropna().unique()
    )

    print("")
    print("Matching selected dates found:")
    for d in available_dates[:100]:
        print(f" - {d}")

    if len(available_dates) > 100:
        print(f"... plus {len(available_dates) - 100} more dates")

    print("")
    print("Sample matching rows:")
    sample_cols = [
        "Hub",
        "Product",
        "TradeDateRaw",
        "StripRaw",
        "DateHeader",
        "SettlementPrice",
    ]
    print(product_matches[sample_cols].head(25).to_string(index=False))
    print("=================================")
    print("")


def filter_daily_prices(
    df: pd.DataFrame,
    start_date: str,
    end_date: str,
) -> pd.DataFrame:
    """
    Filters to:
    - Product contains NG Swing GDD Futures
    - selected date is within requested range
    - Hub is valid
    - Settlement price is valid
    """

    start = pd.Timestamp(start_date).normalize()
    end = pd.Timestamp(end_date).normalize()

    product_matches = df[
        df["Product"].str.contains(PRODUCT_FILTER, case=False, na=False)
    ].copy()

    if product_matches.empty:
        raise ValueError(
            f"No rows found where Product contains '{PRODUCT_FILTER}'."
        )

    available_dates = sorted(
        d for d in product_matches["DateHeader"].dropna().unique()
    )

    filtered = product_matches[
        product_matches["SelectedDate"].notna()
        & (product_matches["SelectedDate"] >= start)
        & (product_matches["SelectedDate"] <= end)
        & product_matches["Hub"].notna()
        & product_matches["SettlementPrice"].notna()
    ].copy()

    if filtered.empty:
        raise ValueError(
            f"No rows found for Product containing '{PRODUCT_FILTER}' "
            f"between {start_date} and {end_date}.\n\n"
            f"Available matching dates in this file are:\n"
            f"{available_dates}"
        )

    filtered = filtered[["DateHeader", "Hub", "SettlementPrice"]]

    filtered = filtered.drop_duplicates(subset=["DateHeader", "Hub"], keep="last")

    return filtered


def validate_expected_date(prices: pd.DataFrame, expected_date: Optional[str]):
    """
    Verifies that the expected date exists in the extracted data.
    """

    if not expected_date:
        return

    expected = pd.Timestamp(expected_date).strftime("%Y-%m-%d")
    dates = sorted(prices["DateHeader"].dropna().unique())

    if expected not in dates:
        raise ValueError(
            f"Expected date {expected} was not extracted.\n"
            f"Extracted dates were: {dates}"
        )

    matching_rows = prices[prices["DateHeader"] == expected]
    print(f"Validation passed: {len(matching_rows)} rows extracted for expected date {expected}.")


def create_or_load_master(master_file: str):
    """
    Creates a blank master workbook if one does not exist.
    Otherwise loads the existing master workbook.

    Correct layout:
    A1 = Date
    Row 1 = Hubs
    Column A = Dates
    Body = Settlement prices
    """

    master_path = Path(master_file)
    master_path.parent.mkdir(parents=True, exist_ok=True)

    if master_path.exists():
        wb = load_workbook(master_path)
        ws = wb.active
        ws.title = "NG Swing GDD"

        a1 = ws.cell(row=1, column=1).value

        if a1 is None or str(a1).strip() == "":
            ws.cell(row=1, column=1).value = "Date"

        elif str(a1).strip().lower() == "hub":
            raise ValueError(
                "The existing master file appears to use the OLD layout with Hub in A1. "
                "Delete or reset master/master_ng_swing_gdd_v2.xlsx so the script can "
                "recreate the correct layout with A1 = Date."
            )

        elif str(a1).strip().lower() != "date":
            raise ValueError(
                f"The existing master file has A1 = '{a1}'. Expected A1 = Date."
            )

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "NG Swing GDD"
        ws.cell(row=1, column=1).value = "Date"

    return wb, ws


def get_existing_dates(ws):
    """
    Reads existing dates from Column A.
    """

    dates = {}

    for row in range(2, ws.max_row + 1):
        raw_value = ws.cell(row=row, column=1).value
        normalized = normalize_date_header(raw_value)

        if normalized:
            dates[normalized] = row
            ws.cell(row=row, column=1).value = normalized

    return dates


def get_existing_hubs(ws):
    """
    Reads existing hubs from Row 1.
    """

    hubs = {}

    for col in range(2, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value

        if value is not None and str(value).strip() != "":
            hubs[str(value).strip()] = col

    return hubs


def ensure_date_rows(ws, date_headers):
    """
    Adds missing dates down Column A.
    """

    existing_dates = get_existing_dates(ws)
    next_row = ws.max_row + 1 if ws.max_row >= 1 else 2

    if ws.cell(row=1, column=1).value is None:
        ws.cell(row=1, column=1).value = "Date"

    for date_header in sorted(date_headers):
        if date_header not in existing_dates:
            ws.cell(row=next_row, column=1).value = date_header
            existing_dates[date_header] = next_row
            next_row += 1

    return existing_dates


def ensure_hub_columns(ws, hub_names):
    """
    Adds missing hubs across Row 1.
    """

    existing_hubs = get_existing_hubs(ws)
    next_col = ws.max_column + 1 if ws.max_column >= 1 else 2

    for hub in sorted(hub_names):
        if hub not in existing_hubs:
            ws.cell(row=1, column=next_col).value = hub
            existing_hubs[hub] = next_col
            next_col += 1

    return existing_hubs


def update_master(ws, prices: pd.DataFrame):
    """
    Writes prices into the corrected matrix:
    - Dates down Column A
    - Hubs across Row 1
    - Settlement prices in the body
    """

    date_headers = sorted(prices["DateHeader"].dropna().unique())
    hub_names = sorted(prices["Hub"].dropna().unique())

    date_map = ensure_date_rows(ws, date_headers)
    hub_map = ensure_hub_columns(ws, hub_names)

    cells_written = 0

    for _, row in prices.iterrows():
        date_header = row["DateHeader"]
        hub = row["Hub"]
        price = row["SettlementPrice"]

        target_row = date_map[date_header]
        target_col = hub_map[hub]

        ws.cell(row=target_row, column=target_col).value = float(price)
        cells_written += 1

    return cells_written


def sort_master(ws):
    """
    Sorts dates top-to-bottom and hubs left-to-right.
    Rebuilds the matrix while preserving values.
    """

    dates = []
    hubs = []

    for row in range(2, ws.max_row + 1):
        raw_date = ws.cell(row=row, column=1).value
        normalized_date = normalize_date_header(raw_date)

        if normalized_date:
            dates.append(normalized_date)

    for col in range(2, ws.max_column + 1):
        hub = ws.cell(row=1, column=col).value

        if hub is not None and str(hub).strip() != "":
            hubs.append(str(hub).strip())

    dates = sorted(set(dates))
    hubs = sorted(set(hubs))

    data = {}

    for row in range(2, ws.max_row + 1):
        raw_date = ws.cell(row=row, column=1).value
        date = normalize_date_header(raw_date)

        if not date:
            continue

        for col in range(2, ws.max_column + 1):
            hub = ws.cell(row=1, column=col).value

            if hub is None:
                continue

            hub = str(hub).strip()
            value = ws.cell(row=row, column=col).value

            if value is not None:
                data[(date, hub)] = value

    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

    ws.cell(row=1, column=1).value = "Date"

    for col_idx, hub in enumerate(hubs, start=2):
        ws.cell(row=1, column=col_idx).value = hub

    for row_idx, date in enumerate(dates, start=2):
        ws.cell(row=row_idx, column=1).value = date

        for col_idx, hub in enumerate(hubs, start=2):
            ws.cell(row=row_idx, column=col_idx).value = data.get((date, hub))


def format_master(ws):
    """
    Applies basic formatting to the master sheet.
    """

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)

    max_row = ws.max_row
    max_col = ws.max_column

    ws.freeze_panes = "B2"

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(2, max_row + 1):
        date_cell = ws.cell(row=row, column=1)
        date_cell.font = Font(bold=True)
        date_cell.fill = header_fill
        date_cell.alignment = Alignment(horizontal="left", vertical="center")
        date_cell.border = border

    for row in range(2, max_row + 1):
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.number_format = "0.000"
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)

        if col == 1:
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 16

    ws.row_dimensions[1].height = 22

    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def verify_master_has_data(ws, prices: pd.DataFrame):
    """
    Verifies that every extracted date has at least one populated price in the master.
    Fails the workflow if not.
    """

    requested_dates = sorted(prices["DateHeader"].dropna().unique())

    date_rows = get_existing_dates(ws)
    hub_cols = get_existing_hubs(ws)

    for date_header in requested_dates:
        if date_header not in date_rows:
            raise ValueError(
                f"Verification failed: Date {date_header} not found in master Column A."
            )

        row_num = date_rows[date_header]
        populated = 0

        for _, col_num in hub_cols.items():
            value = ws.cell(row=row_num, column=col_num).value
            if value is not None:
                populated += 1

        if populated == 0:
            raise ValueError(
                f"Verification failed: Date {date_header} exists in master, "
                f"but no price cells were populated."
            )

        print(f"Verified {date_header}: {populated} populated hub prices in master.")


def save_audit_files(prices: pd.DataFrame, output_dir: str):
    """
    Saves an audit CSV so you can see exactly what was extracted.
    """

    audit_path = Path(output_dir)
    audit_path.mkdir(parents=True, exist_ok=True)

    extracted_file = audit_path / "last_extracted_ng_swing_gdd.csv"
    prices.to_csv(extracted_file, index=False)

    print(f"Audit extract saved to: {extracted_file}")


def save_master(wb, output_file: str):
    """
    Saves the updated master workbook.
    """

    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def save_master_csv(ws, csv_output_file: str):
    """
    Saves a CSV copy of the visible master sheet.
    This makes it easy to confirm in GitHub that dates/prices were actually added.
    """

    output_path = Path(csv_output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows = []

    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    df = pd.DataFrame(rows)

    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")

    df.to_csv(output_path, index=False, header=False)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Update NG Swing GDD Futures master price file."
    )

    parser.add_argument(
        "--daily",
        default="input/icecleared_latest.xlsx",
        help="Path to daily ICE download file.",
    )

    parser.add_argument(
        "--master",
        default="master/master_ng_swing_gdd_v2.xlsx",
        help="Path to existing master file. Created if it does not exist.",
    )

    parser.add_argument(
        "--output",
        default="master/master_ng_swing_gdd_v2.xlsx",
        help="Path to save updated master file.",
    )

    parser.add_argument(
        "--start",
        default=None,
        help="Start date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--end",
        default=None,
        help="End date in YYYY-MM-DD format.",
    )

    parser.add_argument(
        "--date-field",
        default="strip",
        choices=["strip", "trade_date"],
        help="Which ICE date field to use: strip or trade_date. Default is strip.",
    )

    parser.add_argument(
        "--audit-dir",
        default="audit",
        help="Directory for audit output files.",
    )

    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Only validate the input file and extracted rows. Do not update the master.",
    )

    parser.add_argument(
        "--expected-date",
        default=None,
        help="Optional expected date that must be found in extracted rows.",
    )

    return parser.parse_args()


def main():
    args = parse_args()

    start_date = args.start
    end_date = args.end

    if not start_date:
        start_date = input("Enter start date to pull, YYYY-MM-DD: ").strip()

    if not end_date:
        end_date = input("Enter end date to pull, YYYY-MM-DD: ").strip()

    try:
        pd.Timestamp(start_date)
        pd.Timestamp(end_date)
    except Exception as exc:
        raise ValueError("Start and end dates must be valid YYYY-MM-DD dates.") from exc

    print(f"Reading daily file: {args.daily}")
    daily_df = read_daily_file(args.daily, args.date_field, args.audit_dir)

    print_diagnostics(daily_df)

    print(f"Filtering product containing: {PRODUCT_FILTER}")
    print(f"Requested date range: {start_date} through {end_date}")
    print(f"Date field used: {args.date_field}")

    prices = filter_daily_prices(daily_df, start_date, end_date)
    validate_expected_date(prices, args.expected_date)

    print("")
    print("========== EXTRACTED SUMMARY ==========")
    print(f"Rows extracted: {len(prices)}")
    print(f"Dates found: {prices['DateHeader'].nunique()}")
    print(f"Hubs found: {prices['Hub'].nunique()}")
    print("Extracted dates:")
    for d in sorted(prices["DateHeader"].unique()):
        print(f" - {d}")
    print("=======================================")
    print("")

    save_audit_files(prices, args.audit_dir)

    if args.validate_only:
        print("Validation-only mode complete. Master file was not updated.")
        return

    wb, ws = create_or_load_master(args.master)

    cells_written = update_master(ws, prices)
    print(f"Cells written to master before sort: {cells_written}")

    sort_master(ws)
    format_master(ws)
    verify_master_has_data(ws, prices)

    save_master(wb, args.output)

    csv_output = str(Path(args.output).with_suffix(".csv"))
    save_master_csv(ws, csv_output)

    print(f"Master file updated successfully: {args.output}")
    print(f"CSV master mirror saved successfully: {csv_output}")


if __name__ == "__main__":
    main()
